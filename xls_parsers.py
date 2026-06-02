"""
월간일정표 XLS 파서 — Group / NBT / BIO 회사별
입력: XLS/XLSX bytes
출력: [{'company','date','title'}, ...]
"""
import io, re, datetime

# 노이즈 필터 — 일정 제목이 아닌 셀 내용
_NOISE_PATTERNS = [
    re.compile(r'^[\d\.\s]+$'),                    # 숫자만 (예: "28.0")
    re.compile(r'^[◈※●○◯▶▷■□●※]'),               # 특수문자 시작 (footer 표시)
    re.compile(r'^생산일수'),                      # 생산일수 X일
    re.compile(r'^\d{4}-\d{2}-\d{2}'),             # ISO 날짜 문자열
    re.compile(r'^(일|월|화|수|목|금|토|日|月|火|水|木|金|土)요?일?$'),  # 요일 헤더
    re.compile(r'^\d{1,2}:\d{2}(:\d{2})?$'),       # 시간만 (예: "13:30:00")
    re.compile(r'^\(.*\)$'),                       # 괄호만 둘러싼 텍스트 (연속행, 별도 필터됨)
]
_HOLIDAY_KEYWORDS = ['지방선거', '현충일', '공휴일', '대체', '광복절', '개천절', '한글날', '신정', '설날', '추석']

def _norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip()

def _is_noise(text):
    t = _norm(text)
    if not t or len(t) < 2: return True
    for p in _NOISE_PATTERNS:
        if p.match(t): return True
    return False

def _split_cell(text):
    """셀 내 여러 줄을 항목별로 분리 (들여쓰기 라인은 직전 라인에 병합)"""
    raw_lines = str(text or '').split('\n')
    out = []
    for raw in raw_lines:
        stripped = raw.strip()
        if not stripped: continue
        # 들여쓰기 (앞에 공백) → 이전 라인의 연속
        if out and (raw.startswith((' ', '\t')) or stripped.startswith(('(', '＜', '<'))):
            out[-1] = out[-1] + ' ' + stripped
        else:
            out.append(stripped)
    return out

def _detect_year_month(text):
    m = re.search(r'(\d{4})\s*년\s*(\d{1,2})\s*월', text or '')
    if m: return int(m.group(1)), int(m.group(2))
    return None, None

def _extract_day(v, year, month):
    """셀값에서 해당 월의 일자 추출 (없으면 None)"""
    if isinstance(v, datetime.datetime) and v.year == year and v.month == month:
        return v.day
    if isinstance(v, (int, float)) and v == int(v) and 1 <= v <= 31:
        return int(v)
    return None

def _find_date_rows(rows, year, month):
    """캘린더 그리드의 날짜 행 인덱스 + 컬럼→일 매핑 반환"""
    date_rows = []
    for r, row in enumerate(rows):
        days = []
        for c, v in enumerate(row):
            d = _extract_day(v, year, month)
            if d is not None:
                days.append((c, d))
        if len(days) >= 2:
            # 순증가 검증 (가짜 날짜 행 거름)
            vals = [d for _, d in days]
            if all(vals[i+1] > vals[i] for i in range(len(vals)-1)):
                date_rows.append((r, dict(days)))
    return date_rows

# ────────────────────────────────────────────────
# Group / BIO 공통 — 캘린더 그리드, 각 요일 1열
# 연속행(괄호로 시작) 자동 병합
# ────────────────────────────────────────────────
def _parse_grid_single_col(rows, date_rows, year, month, company):
    items = []
    for i, (r, day_cols) in enumerate(date_rows):
        next_r = date_rows[i+1][0] if i+1 < len(date_rows) else len(rows)
        for c, day in day_cols.items():
            # 한 회사의 한 날짜에 대해 r+1 ~ next_r-1 사이 셀들 순차 처리
            day_items = []
            for cr in range(r+1, next_r):
                text = rows[cr][c] if c < len(rows[cr]) else ''
                if not text or not str(text).strip(): continue
                t_str = str(text)
                # 셀 내 멀티라인 처리
                lines = _split_cell(t_str)
                for line in lines:
                    # 괄호로 시작하거나 앞에 공백 있던 라인 → 직전 항목에 병합
                    is_cont = line.startswith(('(', '＜', '<')) or t_str.startswith((' ', '\t'))
                    if is_cont and day_items:
                        day_items[-1] = day_items[-1] + ' ' + line
                        continue
                    if _is_noise(line): continue
                    day_items.append(line)
            for title in day_items:
                items.append({
                    'company': company,
                    'date': f'{year:04d}-{month:02d}-{day:02d}',
                    'title': title,
                })
    return items

def parse_group(data: bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    ws = None
    for sn in wb.sheet_names():
        s = wb.sheet_by_name(sn)
        if s.nrows >= 5 and s.ncols >= 7:
            ws = s; break
    if not ws: return []
    full_text = ' '.join(str(ws.cell_value(r,c)) for r in range(min(3, ws.nrows)) for c in range(ws.ncols))
    year, month = _detect_year_month(full_text)
    if not year or not month: return []
    rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    return _parse_grid_single_col(rows, _find_date_rows(rows, year, month), year, month, 'Group')

def parse_bio(data: bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    ws = None
    for sn in wb.sheet_names():
        s = wb.sheet_by_name(sn)
        if s.nrows >= 5 and s.ncols >= 7:
            ws = s; break
    if not ws: return []
    full_text = ' '.join(str(ws.cell_value(r,c)) for r in range(min(3, ws.nrows)) for c in range(ws.ncols))
    year, month = _detect_year_month(full_text)
    if not year or not month: return []
    rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    return _parse_grid_single_col(rows, _find_date_rows(rows, year, month), year, month, 'BIO')

# ────────────────────────────────────────────────
# NBT (.xlsx) — 17열, 각 요일 [시간|내용] 2칸
# ────────────────────────────────────────────────
def parse_nbt(data: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = None
    for sn in wb.sheetnames:
        s = wb[sn]
        if s.max_row >= 5 and s.max_column >= 10:
            ws = s; break
    if not ws: return []

    full_text = ''
    for r in range(1, min(4, ws.max_row+1)):
        for c in ws[r]:
            full_text += ' ' + str(c.value or '')
    year, month = _detect_year_month(full_text)
    if not year or not month: return []

    nrows = ws.max_row
    ncols = ws.max_column
    rows = [[ws.cell(row=r+1, column=c+1).value for c in range(ncols)] for r in range(nrows)]
    date_rows = _find_date_rows(rows, year, month)

    items = []
    for i, (r, day_cols) in enumerate(date_rows):
        next_r = date_rows[i+1][0] if i+1 < len(date_rows) else nrows
        for cr in range(r+1, next_r):
            for c, day in day_cols.items():
                # NBT: 날짜 컬럼 c에서 time at col c, title at col c+1
                time_v = rows[cr][c] if c < ncols else None
                title_v = rows[cr][c+1] if c+1 < ncols else None
                title = None
                title_text = _norm(title_v) if title_v else ''
                if title_text and _is_noise(title_text):
                    title_text = ''

                if title_text:
                    # time_v 가 datetime.time / 시간 문자열이면 prefix
                    if isinstance(time_v, datetime.time):
                        title = f'{time_v.strftime("%H%M")} {title_text}'
                    elif time_v and isinstance(time_v, str):
                        m = re.match(r'(\d{1,2}):(\d{2})', str(time_v))
                        if m:
                            title = f'{int(m.group(1)):02d}{m.group(2)} {title_text}'
                        else:
                            title = title_text
                    else:
                        title = title_text
                else:
                    # title_v 비어있는데 time_v에 텍스트만 있으면 (예: "HNC 전시회", "지방선거")
                    if time_v and not isinstance(time_v, (datetime.time, datetime.datetime)):
                        s = _norm(time_v)
                        if s and not _is_noise(s):
                            title = s
                if title:
                    items.append({
                        'company': 'NBT',
                        'date': f'{year:04d}-{month:02d}-{day:02d}',
                        'title': title,
                    })
    return items

# ────────────────────────────────────────────────
# 회사 자동 감지
# ────────────────────────────────────────────────
def parse_auto(data: bytes, filename: str = ''):
    fn = filename.lower()
    if 'nbt' in fn or '엔비티' in filename:
        return parse_nbt(data)
    if 'bio' in fn or '바이오' in filename:
        return parse_bio(data)
    if 'group' in fn or '그룹' in filename:
        return parse_group(data)
    if filename.endswith('.xlsx'):
        return parse_nbt(data)
    return parse_group(data)
